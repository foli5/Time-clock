
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
import sqlite3
from datetime import datetime, timedelta, date
import math
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.secret_key = "cambia_esta_llave_por_una_segura"

DB_PATH = "checador.db"

OFFICE_LAT = 19.4925
OFFICE_LON = -99.2564
RADIUS_METERS = 100
TOLERANCIA_MIN = 10

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            pin TEXT NOT NULL UNIQUE,
            hora_entrada TEXT NOT NULL,
            hora_salida TEXT NOT NULL,
            jornada_horas REAL NOT NULL,
            activo INTEGER DEFAULT 1
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER NOT NULL,
            fecha TEXT NOT NULL,
            hora_entrada_real TEXT,
            lat_entrada REAL,
            lon_entrada REAL,
            dentro_zona_entrada INTEGER,
            hora_salida_real TEXT,
            lat_salida REAL,
            lon_salida REAL,
            dentro_zona_salida INTEGER,
            FOREIGN KEY(employee_id) REFERENCES employees(id)
        )
    """)
    conn.commit()

    c.execute("SELECT COUNT(*) as cnt FROM employees")
    if c.fetchone()["cnt"] == 0:
        ejemplo = [
            ("Empleado Ejemplo 1", "1001", "08:00", "17:00", 8.0),
            ("Empleado Ejemplo 2", "1002", "09:00", "18:00", 8.0),
            ("Empleado Ejemplo 3", "1003", "07:30", "15:30", 8.0),
        ]
        c.executemany(
            "INSERT INTO employees (nombre, pin, hora_entrada, hora_salida, jornada_horas) VALUES (?,?,?,?,?)",
            ejemplo
        )
        conn.commit()
    conn.close()

def haversine(lat1, lon1, lat2, lon2):
    R = 6371000
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlambda/2)**2
    c = 2*math.atan2(math.sqrt(a), math.sqrt(1-a))
    return R * c

def dentro_de_zona(lat, lon):
    if lat is None or lon is None:
        return False
    dist = haversine(lat, lon, OFFICE_LAT, OFFICE_LON)
    return dist <= RADIUS_METERS

@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        pin = request.form.get("pin", "").strip()
        conn = get_db()
        emp = conn.execute("SELECT * FROM employees WHERE pin=? AND activo=1", (pin,)).fetchone()
        conn.close()
        if emp:
            session["employee_id"] = emp["id"]
            session["employee_nombre"] = emp["nombre"]
            return redirect(url_for("checador"))
        else:
            flash("PIN no válido. Verifica con tu administrador.", "error")
    return render_template("login.html")

@app.route("/checador")
def checador():
    if "employee_id" not in session:
        return redirect(url_for("login"))
    return render_template("checador.html", nombre=session["employee_nombre"],
                            office_lat=OFFICE_LAT, office_lon=OFFICE_LON, radius=RADIUS_METERS)

@app.route("/marcar", methods=["POST"])
def marcar():
    if "employee_id" not in session:
        return redirect(url_for("login"))

    tipo = request.form.get("tipo")
    lat = request.form.get("lat")
    lon = request.form.get("lon")

    try:
        lat = float(lat)
        lon = float(lon)
    except (TypeError, ValueError):
        flash("No se pudo obtener tu ubicación. Activa el GPS e intenta de nuevo.", "error")
        return redirect(url_for("checador"))

    dentro = dentro_de_zona(lat, lon)
    if not dentro:
        flash("Estás fuera del rango permitido del corporativo (100 m). No se registró el marcaje.", "error")
        return redirect(url_for("checador"))

    employee_id = session["employee_id"]
    hoy = date.today().isoformat()
    ahora = datetime.now().strftime("%H:%M:%S")

    conn = get_db()
    registro = conn.execute(
        "SELECT * FROM attendance WHERE employee_id=? AND fecha=?", (employee_id, hoy)
    ).fetchone()

    if tipo == "entrada":
        if registro:
            flash("Ya tienes registrada tu entrada de hoy.", "error")
        else:
            conn.execute(
                "INSERT INTO attendance (employee_id, fecha, hora_entrada_real, lat_entrada, lon_entrada, dentro_zona_entrada) VALUES (?,?,?,?,?,?)",
                (employee_id, hoy, ahora, lat, lon, 1)
            )
            conn.commit()
            flash(f"Entrada registrada correctamente a las {ahora}.", "success")
    elif tipo == "salida":
        if not registro:
            flash("No tienes entrada registrada el día de hoy.", "error")
        elif registro["hora_salida_real"]:
            flash("Ya tienes registrada tu salida de hoy.", "error")
        else:
            conn.execute(
                "UPDATE attendance SET hora_salida_real=?, lat_salida=?, lon_salida=?, dentro_zona_salida=? WHERE id=?",
                (ahora, lat, lon, 1, registro["id"])
            )
            conn.commit()
            flash(f"Salida registrada correctamente a las {ahora}.", "success")

    conn.close()
    return redirect(url_for("checador"))

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

ADMIN_PASSWORD = "admin123"  # CAMBIAR antes de usar en producción

@app.route("/admin", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        pwd = request.form.get("password")
        if pwd == ADMIN_PASSWORD:
            session["is_admin"] = True
            return redirect(url_for("admin_panel"))
        else:
            flash("Contraseña incorrecta.", "error")
    return render_template("admin_login.html")

def calcular_metricas(emp, registro):
    resultado = {
        "hora_entrada_esperada": emp["hora_entrada"],
        "hora_salida_esperada": emp["hora_salida"],
        "jornada_esperada_horas": emp["jornada_horas"],
        "hora_entrada_real": registro["hora_entrada_real"] if registro else None,
        "hora_salida_real": registro["hora_salida_real"] if registro else None,
        "horas_trabajadas": 0.0,
        "retardo_min": 0,
        "falta": False,
        "tiempo_faltante_horas": 0.0,
    }

    if not registro or not registro["hora_entrada_real"]:
        resultado["falta"] = True
        resultado["tiempo_faltante_horas"] = emp["jornada_horas"]
        return resultado

    hora_ent_esp = datetime.strptime(emp["hora_entrada"], "%H:%M")
    hora_ent_real = datetime.strptime(registro["hora_entrada_real"], "%H:%M:%S")
    hora_ent_real = hora_ent_real.replace(year=hora_ent_esp.year, month=hora_ent_esp.month, day=hora_ent_esp.day)

    diff_min = (hora_ent_real - hora_ent_esp).total_seconds() / 60
    if diff_min > TOLERANCIA_MIN:
        resultado["retardo_min"] = round(diff_min)

    if registro["hora_salida_real"]:
        hora_sal_real = datetime.strptime(registro["hora_salida_real"], "%H:%M:%S")
        hora_sal_real = hora_sal_real.replace(year=hora_ent_esp.year, month=hora_ent_esp.month, day=hora_ent_esp.day)
        horas_trabajadas = (hora_sal_real - hora_ent_real).total_seconds() / 3600
        resultado["horas_trabajadas"] = round(max(horas_trabajadas, 0), 2)
        faltante = emp["jornada_horas"] - resultado["horas_trabajadas"]
        resultado["tiempo_faltante_horas"] = round(max(faltante, 0), 2)
    else:
        resultado["tiempo_faltante_horas"] = emp["jornada_horas"]

    return resultado

@app.route("/admin/panel")
def admin_panel():
    if not session.get("is_admin"):
        return redirect(url_for("admin_login"))

    conn = get_db()
    empleados = conn.execute("SELECT * FROM employees WHERE activo=1").fetchall()

    fecha_filtro = request.args.get("fecha", date.today().isoformat())

    filas = []
    for emp in empleados:
        registro = conn.execute(
            "SELECT * FROM attendance WHERE employee_id=? AND fecha=?", (emp["id"], fecha_filtro)
        ).fetchone()
        metrics = calcular_metricas(emp, registro)
        filas.append({"empleado": emp["nombre"], "fecha": fecha_filtro, **metrics})

    conn.close()
    return render_template("admin_panel.html", filas=filas, fecha_filtro=fecha_filtro)

@app.route("/admin/exportar")
def admin_exportar():
    if not session.get("is_admin"):
        return redirect(url_for("admin_login"))

    fecha_inicio = request.args.get("inicio")
    fecha_fin = request.args.get("fin")
    if not fecha_inicio:
        fecha_inicio = date.today().isoformat()
    if not fecha_fin:
        fecha_fin = date.today().isoformat()

    conn = get_db()
    empleados = conn.execute("SELECT * FROM employees WHERE activo=1").fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Asistencia"

    headers = ["Empleado", "Fecha", "Hora Entrada Esperada", "Hora Entrada Real",
               "Hora Salida Esperada", "Hora Salida Real", "Horas Trabajadas",
               "Retardo (min)", "Falta", "Tiempo Faltante (hrs)"]
    ws.append(headers)
    header_fill = PatternFill(start_color="5F142D", end_color="5F142D", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    d_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
    d_fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()

    d = d_inicio
    while d <= d_fin:
        fecha_str = d.isoformat()
        for emp in empleados:
            registro = conn.execute(
                "SELECT * FROM attendance WHERE employee_id=? AND fecha=?", (emp["id"], fecha_str)
            ).fetchone()
            m = calcular_metricas(emp, registro)
            ws.append([
                emp["nombre"], fecha_str, m["hora_entrada_esperada"], m["hora_entrada_real"] or "-",
                m["hora_salida_esperada"], m["hora_salida_real"] or "-", m["horas_trabajadas"],
                m["retardo_min"], "Sí" if m["falta"] else "No", m["tiempo_faltante_horas"]
            ])
        d += timedelta(days=1)

    conn.close()

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"reporte_asistencia_{fecha_inicio}_a_{fecha_fin}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    return redirect(url_for("admin_login"))

@app.route("/admin/empleados", methods=["GET", "POST"])
def admin_empleados():
    if not session.get("is_admin"):
        return redirect(url_for("admin_login"))

    conn = get_db()
    if request.method == "POST":
        nombre = request.form.get("nombre")
        pin = request.form.get("pin")
        hora_entrada = request.form.get("hora_entrada")
        hora_salida = request.form.get("hora_salida")
        jornada_horas = request.form.get("jornada_horas")
        try:
            conn.execute(
                "INSERT INTO employees (nombre, pin, hora_entrada, hora_salida, jornada_horas) VALUES (?,?,?,?,?)",
                (nombre, pin, hora_entrada, hora_salida, float(jornada_horas))
            )
            conn.commit()
            flash("Empleado agregado correctamente.", "success")
        except sqlite3.IntegrityError:
            flash("Ese PIN ya está en uso, elige otro.", "error")

    empleados = conn.execute("SELECT * FROM employees WHERE activo=1").fetchall()
    conn.close()
    return render_template("admin_empleados.html", empleados=empleados)

@app.route("/admin/empleados/eliminar/<int:emp_id>")
def eliminar_empleado(emp_id):
    if not session.get("is_admin"):
        return redirect(url_for("admin_login"))
    conn = get_db()
    conn.execute("UPDATE employees SET activo=0 WHERE id=?", (emp_id,))
    conn.commit()
    conn.close()
    flash("Empleado eliminado (desactivado).", "success")
    return redirect(url_for("admin_empleados"))

# Inicializa la base de datos al importar el modulo (necesario para gunicorn en Render)
init_db()

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
